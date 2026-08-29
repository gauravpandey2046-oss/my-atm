from openpyxl import Workbook, load_workbook

file = "atm.xlsx"

try:
    wb = load_workbook(file)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active
    ws.append(["Account No", "Name", "Balance"])
    wb.save(file)


def create_account():
    wb = load_workbook(file)
    ws = wb.active

    acc = input("Enter Account Number: ")
    name = input("Enter Name: ")
    balance = float(input("Enter Initial Balance: "))

    ws.append([acc, name, balance])
    wb.save(file)

    print("Account Created Successfully!")


def deposit():
    wb = load_workbook(file)
    ws = wb.active

    acc = input("Enter Account Number: ")
    amount = float(input("Enter Deposit Amount: "))

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == acc:
            row[2].value = row[2].value + amount
            wb.save(file)
            print("Money Deposited Successfully!")
            return

    print("Account Not Found!")


def withdraw():
    wb = load_workbook(file)
    ws = wb.active

    acc = input("Enter Account Number: ")
    amount = float(input("Enter Withdrawal Amount: "))

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == acc:

            if amount <= row[2].value:
                row[2].value = row[2].value - amount
                wb.save(file)
                print("Money Withdrawn Successfully!")
            else:
                print("Insufficient Balance!")

            return

    print("Account Not Found!")


def show_balance():
    wb = load_workbook(file)
    ws = wb.active

    acc = input("Enter Account Number: ")

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == acc:
            print("Name:", row[1].value)
            print("Balance:", row[2].value)
            return

    print("Account Not Found!")


while True:

    print("\n===== ATM SYSTEM =====")
    print("1. Create Account")
    print("2. Deposit")
    print("3. Withdraw")
    print("4. Show Balance")
    print("5. Exit")

    choice = int(input("Enter Choice: "))

    if choice == 1:
        create_account()

    elif choice == 2:
        deposit()

    elif choice == 3:
        withdraw()

    elif choice == 4:
        show_balance()

    elif choice == 5:
        print("Thank You!")
        break

    else:
        print("Wrong Choice!")